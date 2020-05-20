import openpyxl
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from pathlib import Path

# set path to xlsx data sheet                                                                                                                                                                                                                                                                                                                                                                                                                                                                                      
xl = Path(Path.home(), 'Downloads', 'Florida_Data_Revised.xlsx')

# read xslx file                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                   
workbook_obj = openpyxl.load_workbook(xl)


# read the active sheets from file                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
# pgrm_page has the largest amount of data so all data will aggregate into that sheet                                                                                                                                                                                                                                                                                                                                                                                                                              
org_page = workbook_obj['Organization']
pgrm_page = workbook_obj['Program']
contact_page = workbook_obj['Contact']

# list of available organization/program names                                                                                                                                                                                                                                                                                                                                                                                                                                                                     
org_list = []
for row in range(2, 85):
    cell_name = 'A{}'.format(row)
    org_list.append(pgrm_page[cell_name].value)



# iterate through sheet to find orgs missing from list                                                                                                                                                                                                                                                                                                                                                                                                                                                             
max_length = 84
for row in range(2, 51 + 1):
    cell_name = 'A{}'.format(row)
#    if org_page[cell_name].value not in org_list:                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
    org_list.append(org_page[cell_name].value)
    pgrm_page.cell(row=max_length+1, column=1, value=org_page[cell_name].value)
    max_length += 1

for row in range(2, 16 + 1):
    cell_name = 'A{}'.format(row)
    # if contact_page[cell_name].value not in org_list:                                                                                                                                                                                                                                                                                                                                                                                                                                                            
    org_list.append(contact_page[cell_name].value)
    pgrm_page.cell(row=max_length+1, column=1, value=contact_page[cell_name].value)
    max_length += 1


# delete organization ID                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
# pgrm_page.delete_cols(1)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                         

# find orgs from Organization sheet in Program sheet and copy info approtpriatley                                                                                                                                                                                                                                                                                                                                                                                                                                  
max_col = 18

for org_row in range(2, 51 + 1):
    org_cell_name = 'A{}'.format(org_row)
    for pgrm_row in range(2, max_length):
        pgrm_cell_name = 'A{}'.format(pgrm_row)
        if org_page[org_cell_name].value == pgrm_page[pgrm_cell_name].value:
            # better way to do this?                                                                                                                                                                                                                                                                                                                                                                                                                                                                               
            pgrm_page.cell(row=pgrm_row, column=max_col+1, value=org_page['B{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+2, value=org_page['C{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+3, value=org_page['D{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+4, value=org_page['E{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+5, value=org_page['F{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+6, value=org_page['G{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+7, value=org_page['H{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+8, value=org_page['I{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+9, value=org_page['J{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+10, value=org_page['K{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+11, value=org_page['L{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+12, value=org_page['M{}'.format(org_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+13, value=org_page['N{}'.format(org_row)].value)

# set header values for Organization sheet headers                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
# again, better way to do this?                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    
pgrm_page.cell(row=1, column=max_col+1, value=org_page['B1'].value)
pgrm_page.cell(row=1, column=max_col+2, value=org_page['C1'].value)
pgrm_page.cell(row=1, column=max_col+3, value=org_page['D1'].value)
pgrm_page.cell(row=1, column=max_col+4, value=org_page['E1'].value)
pgrm_page.cell(row=1, column=max_col+5, value=org_page['F1'].value)
pgrm_page.cell(row=1, column=max_col+6, value=org_page['G1'].value)
pgrm_page.cell(row=1, column=max_col+7, value=org_page['H1'].value)
pgrm_page.cell(row=1, column=max_col+8, value=org_page['I1'].value)
pgrm_page.cell(row=1, column=max_col+9, value=org_page['J1'].value)
pgrm_page.cell(row=1, column=max_col+10, value=org_page['K1'].value)
pgrm_page.cell(row=1, column=max_col+11, value=org_page['L1'].value)
pgrm_page.cell(row=1, column=max_col+12, value=org_page['M1'].value)
pgrm_page.cell(row=1, column=max_col+13, value=org_page['N1'].value)
max_col += 13

# find orgs from Contact sheet in Program sheet and copy info approtpriatley                                                                                                                                                                                                                                                                                                                                                                                                                                       
for contact_row in range(2, 16 + 1):
    contact_cell_name = 'B{}'.format(contact_row)
    for pgrm_row in range(2, max_length):
        pgrm_cell_name = 'B{}'.format(pgrm_row)
        if contact_page[contact_cell_name].value == pgrm_page[pgrm_cell_name].value:
            # again, better way to do this?                                                                                                                                                                                                                                                                                                                                                                                                                                                                        
            pgrm_page.cell(row=pgrm_row, column=max_col+1, value=contact_page['B{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+1, value=contact_page['C{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+2, value=contact_page['D{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+3, value=contact_page['E{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+4, value=contact_page['F{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+5, value=contact_page['G{}'.format(contact_row)].value)
            pgrm_page.cell(row=pgrm_row, column=max_col+6, value=contact_page['H{}'.format(contact_row)].value)



#set header values for Contact sheet headers                                                                                                                                                                                                                                                                                                                                                                                                                                                                       
pgrm_page.cell(row=1, column=max_col+1, value=contact_page['B1'].value)
pgrm_page.cell(row=1, column=max_col+1, value=contact_page['C1'].value)
pgrm_page.cell(row=1, column=max_col+2, value=contact_page['D1'].value)
pgrm_page.cell(row=1, column=max_col+3, value=contact_page['E1'].value)
pgrm_page.cell(row=1, column=max_col+4, value=contact_page['F1'].value)
pgrm_page.cell(row=1, column=max_col+5, value=contact_page['G1'].value)
pgrm_page.cell(row=1, column=max_col+6, value=contact_page['H1'].value)



# setting default styles                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')

alignment=Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
border = Border(outline=Side(border_style='double',
                             color='FF000000'))


# set styles for every cell in page                                                                                                                                                                                                                                                                                                                                                                                                                                                                                
for row in pgrm_page:
    for cell in row:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

# set styles for every cell in header                                                                                                                                                                                                                                                                                                                                                                                                                                                                              
for cell in pgrm_page['1:1']:
    cell.font=Font(bold=True, size=10.5)
    cell.fill=PatternFill("solid", fgColor="DDDDDD")



pgrm_page.delete_cols(1)

# remove all other sheets                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
workbook_obj.remove(workbook_obj['Contact'])
workbook_obj.remove(workbook_obj['Organization'])

# save changes                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                     
workbook_obj.save(xl)
